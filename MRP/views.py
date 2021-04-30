from django.shortcuts import render, redirect
import openpyxl 
import xlwt
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.models import User, auth
from django.contrib import messages 
from django.contrib.auth.decorators import login_required
from .models import *
from .filters import ChemicalFilter, WeekFilter

# Create your views here.


def HomePage(request):
    return render(request, 'homepage.html')

def Register(request):
    if request.method == "POST" :
        data = request.POST.copy()
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        email = data.get('email')
        password = data.get('password')

        newuser = User()
        newuser.username = email
        newuser.first_name = first_name
        newuser.last_name = last_name 
        newuser.email = email
        newuser.set_password(password)
        newuser.save() 

        return redirect('home-page')


    return render(request, 'register.html')

@login_required
def FileUpLoad(request):
    return render(request, 'fileupload.html')

@login_required
def Material(request):
    chemical = Chemical.objects.all()
    vendor = Vendor.objects.all()
    po = PO.objects.all()
    inv_chem = Inv_Chemical.objects.all()
    if request.method == "POST" :
        data = request.POST.copy()
        deletecheck = data.getlist('check')
        for i in deletecheck :
            Chemical.objects.filter(part_num = i).delete()
            ####ทำให้มันลบ po กับ vendor ด้วยอย่าให้มันลบแค่สารเคมี
        
        
    return render(request, 'material.html', {'chemical':chemical,'vendor':vendor,'po': po,'inv_chem':inv_chem})
def AddMaterial(request):
    if request.method == "POST":
        data = request.POST.copy()
        partnumber = data.get('partnumber')
        chemname = data.get('chemname')
        leadtime = data.get('leadtime')
        stdpack = data.get('stdpack')
        price = data.get('price')
        chemclass = data.get('chemclass')
        onhand = data.get('onhand')
        uom = data.get('uom')
        stdbom = data.get('stdbom')
        #below are not include in db chemical table
        vendorname = data.get('vendorname') #vendor Table
        emailvendor = data.get('emailvendor')  #Vendor Table
        ponum = data.get('ponum') #!!!!ต้องมา design ตาราง PO ที่เป็นแบบ static ว่าจะให้เก็บข้อมูลอะไรบ้าง
        expdate = data.get('expdate') #เข้าไปอยู่ใน inventory
        
        newvendor = Vendor()
        newvendor.vendor_name = vendorname
        newvendor.vendor_cont = emailvendor
        newvendor.save() 

        newpo = PO()
        newpo.po_number = ponum
        newpo.save() 

        newchem = Chemical()
        newchem.part_num = partnumber
        newchem.chem_name = chemname
        newchem.leadtime = leadtime
        newchem.std_packing = stdpack
        newchem.onhand = onhand
        newchem.chem_price = price
        newchem.chem_class = chemclass
        newchem.uom = uom
        newchem.STD_BOM = stdbom
        newchem.vendor_id = newvendor
        newchem.po_number = newpo
        newchem.save()


        new_invchem = Inv_Chemical()
        new_invchem.expired_date = expdate
        new_invchem.save()


        return redirect('material-page')

    return render(request, 'Add_material.html')

def WeekLoad(request):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["CAP"]

        week_n = worksheet['D5'].value[2:4]+"'"+worksheet['D5'].value[5:7]
        week_n1 = worksheet['G5'].value[2:4]+"'"+worksheet['G5'].value[5:7]
        week_n2 = worksheet['J5'].value[2:4]+"'"+worksheet['J5'].value[5:7]
        week_list = [week_n, week_n1, week_n2]

        #TS040/48 week n
        max1 = max(worksheet['C10'].value,worksheet['D10'].value)
        #TS040/48 week n+1
        max2 = max(worksheet['F10'].value,worksheet['G10'].value)
        #TS040/48 week n+2
        max3 = max(worksheet['I10'].value,worksheet['J10'].value)

        #TS056 week n
        max4 = max(worksheet['C11'].value,worksheet['D11'].value)
        #TS056 week n+1
        max5 = max(worksheet['F11'].value,worksheet['G11'].value)
        #TS056 week n+2
        max6 = max(worksheet['I11'].value,worksheet['J11'].value)

        data_list = [max1,max2,max3,max4,max5,max6] 
        package = Package.objects.all() 
        
        x = 0
        for p in package:
            for w in week_list:
                newload = WeekLoading()
                newload.package_id = p
                newload.week = w
                newload.loading = data_list[x]
                x += 1   
                newload.save()

    
        return render(request,'Add_weekloading.html',{'data_list': data_list} )

    return render(request, 'Add_weekloading.html')

#planing ให้ลองrun ดูว่าต้องสั่งมั้ย เอา forecast ที่ได้เเต่ละวีคมาปรับเเล้วลองรันดู
def Planning(request):
    if request.method == "POST":
        searchchem = request.POST['searchchem']
        weekstart = request.POST['weekstart']
        weekend = request.POST['weekend']
        chem_data = Chemical.objects.filter(chem_name=searchchem)
        week_load = WeekLoading.objects.all()
        package = Package.objects.all() 
        actual = Inv_Chemical.objects.all()
        #คิด EOQ/BOQ มาด้วย
        eoq = EoqBoqload.objects.all()
        def myFunc(e):
            return e[3:5]
        each_week = sorted(WeekLoading.objects.values_list('week', flat = True).distinct())
        each_week.sort(key = myFunc)
        week_before_eoq = each_week.copy()
    
        #weekloadingแบบเอา Eoq/boq มาคิดด้วย กรณียังไม่มี 3 week loading  ตอนนี้มีถึง เดือน 9 (วีคที่ 35-39 ) 

        numweek_in_each_month = [4,4,5,4,4,5,4,4,5,4,4,5]
        eoq_week = 0
        for i in range(int(max(eoq.values_list('month', flat = True)))) :
            eoq_week += numweek_in_each_month[i]

        if each_week[-1][3:] <= max(eoq.values_list('year', flat = True))[2:] and int(each_week[-1][:2]) < eoq_week :
            #เดี๋ยวมาเขียนอีกกรณีนึงคือ each_week[-1][3:] < max(eoq.values_list('year', flat = True))[2:]
            if each_week[-1][3:] == max(eoq.values_list('year', flat = True))[2:] :
                for j in range(int(each_week[-1][:2])+1,eoq_week+1) :
                    each_week.append(str(j)+"'"+each_week[-1][3:])

        each_week_after_search = []
     
        for w in range(len(each_week)) : #เอาเเค่วีคที่ search
            if each_week[w] == weekstart :
                start = w 
            if each_week[w] == weekend :
                end = w
        each_week_after_search = each_week[start:end+1]
     
        #actual usage in this chemical (ต้องหาก่อนว่าวีคนั้นอยู่เดือนไร,ปีไร เเล้วเอา actual มาหารจำนวนวีค)
        numweek_in_each_month = [4,4,5,4,4,5,4,4,5,4,4,5] 
        actual_each_week = []
        for ac in each_week: #22'20
            week = int(ac[:2])
            for n in range(len(numweek_in_each_month)) :
                week -= numweek_in_each_month[n]
                if week == 0 :
                    #print(n+1)
                    break
                if week < 0 :
                    #print(n+1)
                    break
            act = actual.filter(year = '20'+ac[3:], month = n+1, part_num=Chemical.objects.get(chem_name=searchchem).part_num, chem_isin=False).values_list('chem_amount', flat=True)
            if len(act) == 0 :
                actual_each_week.append("")
            else :
                actual_each_week.append(int(act[0]/numweek_in_each_month[n]))
        actual_after_search = actual_each_week[start:end+1]
      
        #forecast_adjust
        month = []
        year = []
        balance = []
        onhand = [2880]

        #check ว่า week นั้นอยู่เดือนไหน
        for w in each_week:
            if ('20'+w[3:]) not in year :
                year.append('20'+w[3:])
        
        #adjust forecast by week
        adj_forecast = []
        for i in each_week :
            wi = int(i[:2])
            yi = i[3:]
            for n in range(len(numweek_in_each_month)):
                wi -= numweek_in_each_month[n]
                if wi == 0 :
                    #print(n+1)
                    break
                if wi < 0 :
                    #print(n+1)
                    break
            adj = []
            for y in year :
                if y[2:] < yi :
                    ac = actual.filter(year = y, month = n+1, part_num= Chemical.objects.get(chem_name=searchchem).part_num, chem_isin=False).values_list('chem_amount', flat=True)
                    sumload_thischem = 0
                    for p in package : 
                        sumload_thischem += week_load.filter(week = i[:2]+"'"+y[2:]).filter(package_id = p).values_list('loading', flat = True).last()
                    diff = (ac[0]/numweek_in_each_month[n])-(sumload_thischem * Chemical.objects.get(chem_name=searchchem).STD_BOM)
                    adj.append(diff)
            
            if adj != [] :
                if int(i[:2]) > int(week_before_eoq[-1][:2]) :
                    adj_forecast.append(int(((eoq.filter(year= y, month = n+1).values_list('loading',flat=True)[0]*Chemical.objects.get(chem_name=searchchem).STD_BOM)/numweek_in_each_month[n])+(sum(adj)/len(adj))))
                else :
                    sumload = 0
                    for p in package :
                        sumload += week_load.filter(week = i[:2]+"'"+yi).filter(package_id = p).values_list('loading', flat = True).last()
                    adj_forecast.append(int((sumload * Chemical.objects.get(chem_name=searchchem).STD_BOM)+(sum(adj)/len(adj))))
                
            if yi == year[0][2:] :
                sumload_thischem = 0
                for p in package : 
                    sumload_thischem += week_load.filter(week = i[:2]+"'"+yi).filter(package_id = p).values_list('loading', flat = True).last()
                adj_forecast.append(int(sumload_thischem * Chemical.objects.get(chem_name=searchchem).STD_BOM))
        adj_forecast_after_search = adj_forecast[start:end+1]

        
        #balance = onhand - usage + order recieve
        for x in range(len(adj_forecast)):
            if actual_each_week[x] != "" and len(actual_each_week) >= x :
                if balance == [] :
                    bal = onhand[0] - actual_each_week[x]
                    balance.append(bal)
                else :
                    bal = balance[-1] - actual_each_week[x]
                    balance.append(bal)
            else :
                if balance == [] :
                    bal = onhand[0] - adj_forecast[x]
                    balance.append(bal)
                else :
                    bal = balance[-1] - adj_forecast[x]
                    balance.append(bal)
        balance_after_search = balance[start:end+1]
        #print(balance)

        #check ว่า shortage มั้ย
        order_release = [] #อาจจะให้เก็บ ความยาวเท่ากับ each_week แล้วข้างในเป็น "" หมดเลย
        for i in range(len(each_week)):
            order_release.append("")
        order_receive = []
        for j in range(len(each_week)):
            order_receive.append("")
        week_policy1 = [1,5,9,14,18,22,27,31,35,40,44,48,1] #กรณีที่สั่งสัปดาห์ที่1
        week_policy2 = [2,6,10,15,19,23,28,32,36,41,45,49,2] #กรณีที่สั่งสัปดาห์ที่2

        #inventory position = balance + plan receive ของ 7 วีคข้างหน้า
        inv_pos = balance.copy()

        def Calsum(s) : 
            total = 0
            for i in s :
                if i != "" :
                    total += int(i)
            return total
        for i in range(0,len(each_week)-1):
            fake_bal = balance[i]
            for j in range(1,8) :
                if i+j == 54 :
                    break
                if actual_each_week[i+j] != "" :
                    fake_bal -= actual_each_week[i+j]
                    if fake_bal <= 0  :
                        #หา MAD วีคที่มีการ short-7
                        if i+j >= 7 :
                            total_abs_error = []
                            for m in range(i+j-7+1) : #ex.22
                                if actual_each_week[m] != "" : 
                                    total_abs_error.append(abs(actual_each_week[m]-adj_forecast[m]))
                            mad = sum(total_abs_error)/len(total_abs_error)
                            ss = int(1.6 * 1.65 * mad)
                            order_receive[i+j] = sum(adj_forecast[i+j+1:i+j+6]) + ss #SS = 1.6*1.65*MAD  MAD = ABS error เฉลี่ยของวีคก่อนหน้าทั้งหมดรวมตัวมันเอง 
                        else :
                            order_receive[i+j] = sum(adj_forecast[i+j+1:i+j+6])
                        
                        #ทำให้ order release ตรงกับ policy เรา
                        if int(each_week[i+j][:2])-7 in week_policy1  :
                            order_release[i+j-7] = order_receive[i+j] 
                        elif int(each_week[i+j][:2])-7  in week_policy2 :
                            order_release[i+j-7] = order_receive[i+j]
                        else :
                            for n in range(len(week_policy2)) :
                                if int(each_week[i+j][:2])-7  < week_policy2[n] : #ex. 33  สั่งวีค 23 i+j == 11
                                    order_release[i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1])] = order_receive[i+j]
                                    break
                        balance[i+j] = balance[i+j-1] + order_receive[i+j] - actual_each_week[i+j]
                        for b in range(len(balance[i+j+1:])):
                            if actual_each_week[b+1+j+i] != "" :
                                balance[i+j+1+b] = balance[i+j+b] - actual_each_week[i+j+1+b]
                            else :
                                balance[i+j+1+b] = balance[i+j+b] - adj_forecast[i+j+1+b]
                       
                        fake_bal = balance[i+j]
                else :
                    fake_bal -= adj_forecast[i+j]
                    if fake_bal <= 0 :
                        #หา MAD วีคที่มีการ short-7
                        total_abs_error = []
                        for m in range(i+j-7+1) : #ex.22
                            if actual_each_week[m] != "" : 
                                total_abs_error.append(abs(actual_each_week[m]-adj_forecast[m]))
                        mad = sum(total_abs_error)/len(total_abs_error)
                        ss = int(1.6 * 1.65 * mad)
                        order_receive[i+j] = sum(adj_forecast[i+j+1:i+j+6]) + ss
                        #ทำให้ order release ตรงกับ policy เรา
                        if int(each_week[i+j][:2])-7 in week_policy1  :
                            order_release[i+j-7] = order_receive[i+j] 
                        elif int(each_week[i+j][:2])-7  in week_policy2 :
                            order_release[i+j-7] = order_receive[i+j]
                        else :
                            for n in range(len(week_policy2)) :
                                if int(each_week[i+j][:2])-7  < week_policy2[n] : #ex. 33  สั่งวีค 23 i+j == 11
                                    order_release[i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1])] = order_receive[i+j]
                                    break
                        balance[i+j] = balance[i+j-1] + order_receive[i+j] - adj_forecast[i+j]
                        for b in range(len(balance[i+j+1:])):
                            if actual_each_week[b+1+j+i] != "" :
                                balance[i+j+1+b] = balance[i+j+b] - actual_each_week[i+j+1+b]
                            else :
                                balance[i+j+1+b] = balance[i+j+b] - adj_forecast[i+j+1+b]

                        fake_bal = balance[i+j]
            inv_pos[i] = Calsum(order_receive[i+1:i+8]) + balance[i] 
        
            if inv_pos[i-1] < adj_forecast[i] and i >=1 :  #ให้สั่งถ้า inv pos[n+6] < fore usage[n+7]
                print("KUAY",inv_pos[i-1],adj_forecast[i] )

        #print(balance)
        #print(order_receive)
        #print(order_release) 
        
        #ถ้ามี order ให้ขึ้นไปที่หน้า dashboard ด้วย 
        #remark ถ้าไม่ search ก็จะไม่รู้เลยว่าสารไหนshort????

    
    
        return render(request, 'Planning_table.html', {'chem_data':chem_data,'inv_pos': inv_pos,'order_release':order_release,'order_receive': order_receive, 'start':start,'balance':balance, 'end':end, 'each_week_after_search':each_week_after_search, 'adj_forecast_after_search':adj_forecast_after_search, 'actual_after_search':actual_after_search, 'onhand':onhand, 'balance':balance})

            
    return render(request, 'Planning_table.html')

def ActualUsage(request):
    if request.method == "POST":
        y = request.POST['year']
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        worksheet = wb["USAGE"]

        num = 0
        for row in range(3,4):
            for col in range(1,16):
                if worksheet.cell(row,col).value == 0  :
                    num = col
                    break
                else :
                    num = 16
        
        partnum = []
        for row in range(3,24):
            for col in range(1,num):
                if col != 2 and col != 3 :
                    if col == 1 :
                        partnum.append(worksheet.cell(row,col).value)
                        #print(part_num)
                    else : 
                        chemamount = float(worksheet.cell(row,col).value)
                        c = Inv_Chemical(year = y, month = str(col-3), chem_isin = False, chem_amount= chemamount, part_num_id = partnum[-1])
                        c.save()
        #check ก่อนว่า file ชื่อเดียวกันไหม ถ้าชื่อเดียวกันก็ขึ้นว่าเคยอัพโหลดไฟล์นี้เเล้ว    
                        #print(amount)
        #Inv_Chemical.objects.all().delete()
        print(num)
        
    return render(request, 'Add_actualusage.html')

def EoqBoq(request):
    if request.method == "POST":
        allpackage = Package.objects.all()
        data = request.POST.copy()
        y = data.get('year')
        m = data.get('m')
        q = data.get('q')
        lga = data.get('lga')
        ts048 = data.get('ts048')
        ts056 = data.get('ts056')
    

        if request.POST['q'] == '' :
            eoqboq = EoqBoqload(year = y, month = m, loading = 40/48 * int(lga) + int(ts048) + 56/48 * int(ts056) )
            eoqboq.save()
        
        else : 
            end = (int(q)*3)+1
            for i in range(end-3,end):
                eoqboq = EoqBoqload(year = y, month = i , loading = (40/48 * int(lga) + int(ts048) + 56/48 * int(ts056))//3 )
                eoqboq.save()
              
            #ถ้ามันรวมมาเป็น quater เราเเยกครึ่งๆเลยได้ใช่มั้ยว่ามาจาก TS048 ครึ่งนึง TS056 ครึ่งนึง
        return render(request, 'Add_eoqboq.html')
            
    
    return render(request, 'Add_eoqboq.html')

def DashBoard(request):
    chemical = Chemical.objects.all()
    week_load = WeekLoading.objects.all()
    package = Package.objects.all() 
    actual = Inv_Chemical.objects.all()
    #คิด EOQ/BOQ มาด้วย
    eoq = EoqBoqload.objects.all()

    def FindMonth(ww):  #ช่วยหาว่าวีคนั้นอยู่เดือนอะไร ww = XX'XX
        numweek = [4,4,5,4,4,5,4,4,5,4,4,5]
        w = int(ww[:2])
        for n in range(len(numweek_in_each_month)) :
            w -= numweek[n]
            if w == 0 :
                return n+1
            if w < 0 :
                return n+1
        
    def myFunc(e):
        return e[3:5]
    each_week = sorted(WeekLoading.objects.values_list('week', flat = True).distinct())
    each_week.sort(key = myFunc)
    week_before_eoq = each_week.copy()
    
    #weekloadingแบบเอา Eoq/boq มาคิดด้วย กรณียังไม่มี 3 week loading  ตอนนี้มีถึง เดือน 9 (วีคที่ 35-39 ) 

    numweek_in_each_month = [4,4,5,4,4,5,4,4,5,4,4,5]
    eoq_week = 0
    for i in range(int(max(eoq.values_list('month', flat = True)))) :
        eoq_week += numweek_in_each_month[i]

    if each_week[-1][3:] <= max(eoq.values_list('year', flat = True))[2:] and int(each_week[-1][:2]) < eoq_week :
        #เดี๋ยวมาเขียนอีกกรณีนึงคือ each_week[-1][3:] < max(eoq.values_list('year', flat = True))[2:]
        if each_week[-1][3:] == max(eoq.values_list('year', flat = True))[2:] :
            for j in range(int(each_week[-1][:2])+1,eoq_week+1) :
                each_week.append(str(j)+"'"+each_week[-1][3:])
 
    order = [] #จะเอาไว้เก็บเพื่อโชว์หน้า dashboard ว่าต้องสั่งวันไหน  
    #ต้องรันสารเเต่ละตัวเพื่อดูว่ามีตัวไหน short มั้ย
    #300084 #สมมติเราอยู่วีคที่ n คือวีคที่ 29'20 (อิงมาจากวีคที่มีการอัพโหลดไฟล์ forecast usage เช่น ถ้าอัพของวีคที่ 29,30,31 แสดงว่าตอนนี้เราอยู่วีคที่ 29)
    # ทำให้เวลาเราบันทึก order release เราจะบันทึกถึงแค่วีคที่ 29+7'20 
    for chem in chemical :
        if chem.part_num == "30000084" : #รันสมมติแค่ตัวเดียว
                numweek_in_each_month = [4,4,5,4,4,5,4,4,5,4,4,5] 
                actual_each_week = []
                for ac in each_week: 
                    week = int(ac[:2])
                    for n in range(len(numweek_in_each_month)) :
                        week -= numweek_in_each_month[n]
                        if week == 0 :
                            break
                        if week < 0 :
                            break
                    act = actual.filter(year = '20'+ac[3:], month = n+1, part_num=chem.part_num, chem_isin=False).values_list('chem_amount', flat=True)
                    if len(act) == 0 :
                        actual_each_week.append("")
                    else :
                        actual_each_week.append(int(act[0]/numweek_in_each_month[n]))
                
                #forecast_adjust
                month = []
                year = []
                balance = []
                onhand = [2880]

                #check ว่า week นั้นอยู่เดือนไหน
                for w in each_week:
                    if ('20'+w[3:]) not in year :
                        year.append('20'+w[3:])

                #adjust forecast by week
                adj_forecast = []
                for i in each_week :
                    wi = int(i[:2])
                    yi = i[3:]
                    for n in range(len(numweek_in_each_month)):
                        wi -= numweek_in_each_month[n]
                        if wi == 0 :
                            break
                        if wi < 0 :
                            break
                    adj = []
                    for y in year :
                        if y[2:] < yi :
                            ac = actual.filter(year = y, month = n+1, part_num= chem.part_num, chem_isin=False).values_list('chem_amount', flat=True)
                            sumload_thischem = 0
                            for p in package : 
                                sumload_thischem += week_load.filter(week = i[:2]+"'"+y[2:]).filter(package_id = p).values_list('loading', flat = True).last()
                            diff = (ac[0]/numweek_in_each_month[n])-(sumload_thischem * chem.STD_BOM)
                            adj.append(diff)
                    
                    if adj != [] :
                        if int(i[:2]) > int(week_before_eoq[-1][:2]) :
                            adj_forecast.append(int(((eoq.filter(year= y, month = n+1).values_list('loading',flat=True)[0]*chem.STD_BOM)/numweek_in_each_month[n])+(sum(adj)/len(adj))))
                        else :
                            sumload = 0
                            for p in package :
                                sumload += week_load.filter(week = i[:2]+"'"+yi).filter(package_id = p).values_list('loading', flat = True).last()
                            adj_forecast.append(int((sumload * chem.STD_BOM)+(sum(adj)/len(adj))))

                    if yi == year[0][2:] :
                        sumload_thischem = 0
                        for p in package : 
                            sumload_thischem += week_load.filter(week = i[:2]+"'"+yi).filter(package_id = p).values_list('loading', flat = True).last()
                        adj_forecast.append(int(sumload_thischem * chem.STD_BOM))
                
                #balance = onhand - usage + order recieve
                for x in range(len(adj_forecast)):
                    if actual_each_week[x] != "" and len(actual_each_week) >= x :
                        if balance == [] :
                            bal = onhand[0] - actual_each_week[x]
                            balance.append(bal)
                        else :
                            bal = balance[-1] - actual_each_week[x]
                            balance.append(bal)
                    else :
                        if balance == [] :
                            bal = onhand[0] - adj_forecast[x]
                            balance.append(bal)
                        else :
                            bal = balance[-1] - adj_forecast[x]
                            balance.append(bal)

                #check ว่า shortage มั้ย
                order_release = [] #อาจจะให้เก็บ ความยาวเท่ากับ each_week แล้วข้างในเป็น "" หมดเลย
                for i in range(len(each_week)):
                    order_release.append("")
                order_receive = []
                for j in range(len(each_week)):
                    order_receive.append("")
                week_policy1 = [1,5,9,14,18,22,27,31,35,40,44,48,1] #กรณีที่สั่งสัปดาห์ที่1
                week_policy2 = [2,6,10,15,19,23,28,32,36,41,45,49,2] #กรณีที่สั่งสัปดาห์ที่2

                #inventory position = balance + plan receive ของ 7 วีคข้างหน้า
                inv_pos = balance.copy()

                def Calsum(s) : 
                    total = 0
                    for i in s :
                        if i != "" :
                            total += int(i)
                    return total

                
                for i in range(0,len(each_week)-1):
                    fake_bal = balance[i]
                    for j in range(1,8) :
                        if i+j == 54 :
                            break
                        if actual_each_week[i+j] != "" :
                            fake_bal -= actual_each_week[i+j]
                            if fake_bal <= 0  :
                                if i+j >= 7 :
                                    total_abs_error = []
                                    for m in range(i+j-7+1) : #ex.22
                                        if actual_each_week[m] != "" : 
                                            total_abs_error.append(abs(actual_each_week[m]-adj_forecast[m]))
                                    print("try", total_abs_error, i,j)
                                    mad = sum(total_abs_error)/len(total_abs_error)
                                    ss = int(1.6 * 1.65 * mad)
                                    order_receive[i+j] = sum(adj_forecast[i+j+1:i+j+6]) + ss #SS = 1.6*1.65*MAD  MAD = ABS error เฉลี่ยของวีคก่อนหน้าทั้งหมดรวมตัวมันเอง 
                                else :
                                    order_receive[i+j] = sum(adj_forecast[i+j+1:i+j+6])
                                
                            
                                #ทำให้ order release ตรงกับ policy เรา
                                if int(each_week[i+j][:2])-7 in week_policy1  :
                                    order_release[i+j-7] = order_receive[i+j] 
                                    #ตอนนี้อยู่วีคที่ 29 รันเเล้วมันต้องสั่งวีคที่ 31 ก็ order.append(วีคที่กับปีที่ต้อง release) แล้วให้เก็บ order receive ใน database ต้องแปลงเป็นเดือนว่าอยู่เดือนไหนก่อนแล้วค่อยเก็บ
                                    if int(week_before_eoq[-3][:2]) < int(each_week[i+j][:2])-7 < int(week_before_eoq[-3][:2])+3 and each_week[i+j][3:] == year[-1][:2] :
                                        order.append(str(int(each_week[i+j][:2])-7)+each_week[i+j][2:])
                                        order_quan.append(order_receive[i+j])
                                elif int(each_week[i+j][:2])-7  in week_policy2 :
                                    order_release[i+j-7] = order_receive[i+j]
                                    if int(week_before_eoq[-3][:2]) < int(each_week[i+j][:2])-7 < int(week_before_eoq[-3][:2])+3 and each_week[i+j][3:] == year[-1][:2]:
                                        order.append(str(int(each_week[i+j][:2])-7)+each_week[i+j][2:])
                                        order_quan.append(order_receive[i+j])
                                else :
                                    for n in range(len(week_policy2)) :
                                        if int(each_week[i+j][:2])-7  < week_policy2[n] : #ex. 33  สั่งวีค 23 i+j == 11
                                            order_release[i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1])] = order_receive[i+j]
                                            if int(week_before_eoq[-3][:2]) < i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1]) < int(week_before_eoq[-3][:2])+3 and each_week[i+j][3:] == year[-1][:2] :
                                                order.append(str(i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1]))+each_week[i+j][2:])
                                                order_quan.append(order_receive[i+j])
                                            break
                                balance[i+j] = balance[i+j-1] + order_receive[i+j] - actual_each_week[i+j]
                                for b in range(len(balance[i+j+1:])):
                                    if actual_each_week[b+1+j+i] != "" :
                                        balance[i+j+1+b] = balance[i+j+b] - actual_each_week[i+j+1+b]
                                    else :
                                        balance[i+j+1+b] = balance[i+j+b] - adj_forecast[i+j+1+b]
                            
                                fake_bal = balance[i+j]
                        else :
                            fake_bal -= adj_forecast[i+j]
                            if fake_bal <= 0 :
                                #หา MAD วีคที่มีการ short-7
                                total_abs_error = []
                                for m in range(i+j-7+1) : #ex.22
                                    if actual_each_week[m] != "" : 
                                        total_abs_error.append(abs(actual_each_week[m]-adj_forecast[m]))
                                mad = sum(total_abs_error)/len(total_abs_error)
                                ss = int(1.6 * 1.65 * mad)
                                order_receive[i+j] = sum(adj_forecast[i+j+1:i+j+6]) + ss
                                #ทำให้ order release ตรงกับ policy เรา
                                if int(each_week[i+j][:2])-7 in week_policy1  :
                                    order_release[i+j-7] = order_receive[i+j] 
                                    #ตอนนี้อยู่วีคที่ 29 รันเเล้วมันต้องสั่งวีคที่ 31 ก็ order.append(วีคที่กับปีที่ต้อง release) แล้วให้เก็บ order receive ใน database ต้องแปลงเป็นเดือนว่าอยู่เดือนไหนก่อนแล้วค่อยเก็บ
                                    if int(week_before_eoq[-3][:2]) < int(each_week[i+j][:2])-7 < int(week_before_eoq[-3][:2])+3 and each_week[i+j][3:] == year[-1][:2] :
                                        order.append([chem.part_num,str(int(each_week[i+j][:2])-7)+each_week[i+j][2:],order_receive[i+j]]) #ต้องเปลี่ยนวิธีเก็บ กรณีมันมีหลายสารเคมี อาจจะเก็บเป็น listซ้อนlist [['เก็บวีคกับปีที่ต้องสั่ง','เก็บpartnumของสารเคมี'],[],...]
                                        #order_quan.append(order_receive[i+j])
                                        #เก็บใน database --> order receive, show status
                                        if Inv_Chemical.objects.filter(year = "20"+each_week[i+j][3:], month = FindMonth(each_week[i+j]), chem_isin=True, chem_amount=order_receive[i+j], part_num_id = chem.part_num).exists():
                                            print("exist")
                                        else :
                                            receive = Inv_Chemical(year = "20"+each_week[i+j][3:], month = FindMonth(each_week[i+j]), chem_isin=True, chem_amount=order_receive[i+j], part_num_id = chem.part_num)
                                            print("not exist")
                                            receive.save()
                                            statuschem = Status_Chem(chem_status= "shortage", listchem= chem)
                                            statuschem.save()

                                elif int(each_week[i+j][:2])-7  in week_policy2 :
                                    order_release[i+j-7] = order_receive[i+j]
                                    if int(week_before_eoq[-3][:2]) < int(each_week[i+j][:2])-7 < int(week_before_eoq[-3][:2])+3 and each_week[i+j][3:] == year[-1][:2]:
                                        order.append(str(int(each_week[i+j][:2])-7)+each_week[i+j][2:])
                                        order_quan.append(order_receive[i+j])
                                else :
                                    for n in range(len(week_policy2)) :
                                        if int(each_week[i+j][:2])-7  < week_policy2[n] : #ex. 33  สั่งวีค 23 i+j == 11
                                            order_release[i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1])] = order_receive[i+j]
                                            if int(week_before_eoq[-3][:2]) < i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1]) < int(week_before_eoq[-3][:2])+3 and each_week[i+j][3:] == year[-1][:2] :
                                                order.append(str(i+j-7-((int(each_week[i+j][:2])-7)-week_policy2[n-1]))+each_week[i+j][2:])
                                                order_quan.append(order_receive[i+j])
                                            break
                                balance[i+j] = balance[i+j-1] + order_receive[i+j] - adj_forecast[i+j]
                                for b in range(len(balance[i+j+1:])):
                                    if actual_each_week[b+1+j+i] != "" :
                                        balance[i+j+1+b] = balance[i+j+b] - actual_each_week[i+j+1+b]
                                    else :
                                        balance[i+j+1+b] = balance[i+j+b] - adj_forecast[i+j+1+b]

                                fake_bal = balance[i+j]
                    inv_pos[i] = Calsum(order_receive[i+1:i+8]) + balance[i] 
                
                    if inv_pos[i-1] < adj_forecast[i] and i >=1 :  #ให้สั่งถ้า inv pos[n+6] < fore usage[n+7]
                        print("KUAY",inv_pos[i-1],adj_forecast[i] )
    
    #ดึง status กับ order receive, order quan จาก databaseมา show ในหน้า dashboard
    #ดึง status
    shortage = Status_Chem.objects.filter(chem_status="shortage").values_list('listchem', flat = True)
    print(shortage)

    #ดึงมาโชว์หน้า database สิ่งที่ต้องโชว์มี partnum, name, vendorname, eoq, order release, order receive
    my_order = {} #dict เก็บข้อมูลทั้งหมดเพื่อเอามาโชว์หน้า dashboard
    
    for o in order :
        if o[0] not in my_order :
            my_order[o[0]] = o[1:]

    print("TEST1",order)
    print(my_order)
    
    return render(request,'dashboard.html', {'chemical':chemical, 'shortage':shortage, 'my_order':my_order})

def UpdateMaterial(request, part_num):
    if request.method == "POST":
        data = request.POST.copy()
        partnumber = data.get('partnumber')
        chemname = data.get('chemname')
        leadtime = data.get('leadtime')
        stdpack = data.get('stdpack')
        price = data.get('price')
        chemclass = data.get('chemclass')
        onhand = data.get('onhand')
        uom = data.get('uom')
        stdbom = data.get('stdbom')
        vendorname = data.get('vendorname')
        emailvendor = data.get('emailvendor') 
        ponum = data.get('ponum') 
        expdate = data.get('expdate') 
        vendor_update = Vendor(vendor_name=vendorname, vendor_cont=emailvendor)
        vendor_update.save()
        po_update = PO(po_number=ponum)
        po_update.save()
        chem_update = Chemical( part_num = partnumber, chem_name= chemname, leadtime= leadtime, 
                                std_packing= stdpack, chem_price= price, chem_class=chemclass, 
                                onhand= onhand, uom=uom, STD_BOM= stdbom, vendor_id=vendor_update, po_number = po_update)
        chem_update.save()
        return redirect('material-page')
        
    getchem = Chemical.objects.get(part_num = part_num)
    return render(request, 'updatemat.html', {'getchem':getchem})

def ExportPage(request) :
    status_shortage = Status_Chem.objects.filter(chem_status = "shortage").values_list('listchem', flat = True).distinct()
    chem = Chemical.objects.all()
    quan = Inv_Chemical.objects.all()
    mydict = {} #dict ที่มี value เป็น list จะได้รู้ว่า po นี่มีสารเคมีตัวไหนบ้าง
    partnum = {} #เก็บ key คือ partnum value คือ vendorname, chemname, price, quantiy, amount เป็นเท่าไหร่

    for i in status_shortage :
        po = chem.get(part_num=i).po_number
        if po in mydict :
            mydict[po].append(i)
        else :
            mydict[po] = [i]
        partnum[i] = [  chem.get(part_num=i).vendor_id.vendor_name, chem.get(part_num=i).chem_name, chem.get(part_num=i).chem_price, 
                        quan.filter(part_num=i, chem_isin=True).values_list('chem_amount', flat=True).last(), 
                        quan.filter(part_num=i, chem_isin=True).values_list('chem_amount', flat=True).last() * chem.get(part_num=i).chem_price ]

    print(mydict)
    print(partnum)
    return render(request, 'exportpage.html', {'mydict' : mydict, 'chem':chem, 'partnum':partnum})

def ExportFile(request):
    status_shortage = Status_Chem.objects.filter(chem_status = "shortage").values_list('listchem', flat = True).distinct()
    print(status_shortage)
    chem = Chemical.objects.all()
    quan = Inv_Chemical.objects.all()
    mydict = {} #dict ที่มี value เป็น list
    for i in status_shortage :
        po = chem.get(part_num=i).po_number
        if po in mydict :
            mydict[po].append(i)
        else :
            mydict[po] = [i]
    response = HttpResponse(content_type = 'application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=PO' + '.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('PO')
    rows = len(status_shortage)
    column = ['Part Number', 'Vendor Name', 'Chemical Name', 'Unit Price', 'Quantity', 'Amount']
    for col in range(len(column)) :
        ws.write(0,col,column[col])

    for row in range(1,rows+1):
        for col in range(len(column)):
            if col == 0 :
                ws.write(row,col,status_shortage[row-1])
            elif col == 1 :
                ws.write(row,col,chem.get(part_num=status_shortage[row-1]).vendor_id.vendor_name)
            elif col == 2 :
                ws.write(row,col,chem.get(part_num=status_shortage[row-1]).chem_name)
            elif col == 3 :
                ws.write(row,col,chem.get(part_num=status_shortage[row-1]).chem_price)
            elif col == 4 :
                ws.write(row,col,quan.filter(part_num=status_shortage[row-1], chem_isin=True).values_list('chem_amount', flat=True).last())
            else :
                ws.write(row,col,chem.get(part_num=status_shortage[row-1]).chem_price * quan.filter(part_num=status_shortage[row-1], chem_isin=True).values_list('chem_amount', flat=True).last() )  
    wb.save(response)
    
    #พอกดที่ลิ้งค์ปุ๊ปให้อัพเดทสารตัวนั้นเป็น enough
    for i in status_shortage :
        status = Status_Chem.objects.filter(listchem=chem.get(part_num=i)).update(chem_status = "enough")
     

    return response
    
    

